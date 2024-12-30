import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, List
import json
from io import BytesIO
import base64
from fpdf import FPDF
from openpyxl import Workbook
from datetime import datetime

class EnhancedAnalyzer:
    def __init__(self):
        self.visualization_types = {
            'radar': self._create_radar_chart,
            'timeline': self._create_timeline_chart,
            'distribution': self._create_distribution_chart,
            'heatmap': self._create_heatmap
        }

    def export_report(self, analysis_data: Dict, format: str = 'pdf') -> BytesIO:
        if format == 'pdf':
            return self._export_pdf(analysis_data)
        elif format == 'xlsx':
            return self._export_excel(analysis_data)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _export_pdf(self, analysis_data: Dict) -> BytesIO:
        pdf = FPDF()
        pdf.add_page()
        
        # Add title
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Resume Analysis Report', 0, 1, 'C')
        
        # Add summary
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 10, f"Overall Match Score: {analysis_data['match_score']}%", 0, 1)
        
        # Add visualizations
        for chart_type, chart_data in analysis_data['visualizations'].items():
            img_data = base64.b64decode(chart_data)
            pdf.image(BytesIO(img_data), x=10, y=pdf.get_y(), w=190)
            pdf.ln(100)  # Add space after each chart
        
        output = BytesIO()
        pdf.output(output, 'F')
        output.seek(0)
        return output

    def _export_excel(self, analysis_data: Dict) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Summary"
        
        # Add summary
        ws['A1'] = "Resume Analysis Report"
        ws['A2'] = "Generated at"
        ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A3'] = "Overall Match Score"
        ws['B3'] = f"{analysis_data['match_score']}%"
        
        # Add skills analysis
        ws.create_sheet("Skills Analysis")
        skills_sheet = wb["Skills Analysis"]
        skills_sheet['A1'] = "Matching Skills"
        skills_sheet['B1'] = "Missing Skills"
        
        for i, skill in enumerate(analysis_data['skills_match']['matching_skills'], 2):
            skills_sheet[f'A{i}'] = skill
        for i, skill in enumerate(analysis_data['skills_match']['missing_skills'], 2):
            skills_sheet[f'B{i}'] = skill
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_radar_chart(self, data: Dict) -> str:
        plt.figure(figsize=(8, 8))
        categories = list(data.keys())
        values = list(data.values())
        
        angles = np.linspace(0, 2*np.pi, len(categories), endpoint=False)
        values += values[:1]
        angles += angles[:1]
        
        ax = plt.subplot(111, projection='polar')
        ax.plot(angles, values)
        ax.fill(angles, values, alpha=0.25)
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories)
        
        return self._fig_to_base64()

    def _create_timeline_chart(self, data: List[Dict]) -> str:
        plt.figure(figsize=(10, 6))
        dates = [d['date'] for d in data]
        values = [d['value'] for d in data]
        
        plt.plot(dates, values, marker='o')
        plt.xticks(rotation=45)
        plt.title('Analysis Timeline')
        
        return self._fig_to_base64()

    def _create_distribution_chart(self, data: Dict) -> str:
        plt.figure(figsize=(8, 8))
        labels = list(data.keys())
        sizes = list(data.values())
        
        plt.pie(sizes, labels=labels, autopct='%1.1f%%')
        plt.title('Skills Distribution')
        
        return self._fig_to_base64()

    def _create_heatmap(self, data: List[List[float]], labels: List[str]) -> str:
        plt.figure(figsize=(10, 8))
        sns.heatmap(data, annot=True, xticklabels=labels, yticklabels=labels)
        plt.title('Skills Correlation Matrix')
        
        return self._fig_to_base64()

    def _fig_to_base64(self) -> str:
        """Convert matplotlib figure to base64 string."""
        buf = BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        plt.close()
        buf.seek(0)
        return base64.b64encode(buf.getvalue()).decode('utf-8')

    def create_real_time_analysis(self, resume_content: str, job_description: str) -> Dict:
        """Generate real-time analysis updates."""
        return {
            "progress": 0,
            "current_stage": "Initializing analysis",
            "updates": []
        }